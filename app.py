from __future__ import annotations

from datetime import datetime
from pathlib import Path
from urllib.parse import quote

from flask import Flask, jsonify, render_template, request
from openpyxl import Workbook, load_workbook

BASE_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = BASE_DIR / "data" / "lagerbestand.xlsx"
HEADER_ROW = 3
DATA_START_ROW = 4

app = Flask(__name__)


def to_float(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def format_number(value: float) -> str:
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.2f}".rstrip("0").rstrip(".")


def create_default_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    # Startbeispiel aus der Sperrholz-Uebersicht:
    # nur gelb markierte Staerken, jeweils Sollbestand 2 Stk.
    demo_data = {
        "Sperrholz Pappel": {
            "supplier_email": "lieferant-pappel@example.com",
            "items": [
                ("2520x1870 - 4 mm", 2, 0, "Stk"),
                ("2520x1870 - 6 mm", 2, 0, "Stk"),
                ("2520x1870 - 8 mm", 2, 0, "Stk"),
                ("2520x1870 - 10 mm", 2, 0, "Stk"),
                ("2520x1870 - 12 mm", 2, 0, "Stk"),
                ("2520x1870 - 15 mm", 2, 0, "Stk"),
                ("2520x1870 - 18 mm", 2, 0, "Stk"),
            ],
        },
        "Bausperrholz": {
            "supplier_email": "lieferant-bausperrholz@example.com",
            "items": [
                ("2500x1250 - 9 mm", 2, 0, "Stk"),
                ("2500x1250 - 12 mm", 2, 0, "Stk"),
                ("2500x1250 - 15 mm", 2, 0, "Stk"),
                ("2500x1250 - 18 mm", 2, 0, "Stk"),
            ],
        },
    }

    for group_name, group_data in demo_data.items():
        ws = wb.create_sheet(title=group_name)
        ws["A1"] = "LieferantenEmail"
        ws["B1"] = group_data["supplier_email"]
        ws["A2"] = "Lagergruppe"
        ws["B2"] = group_name
        ws["D1"] = "ZuletztAktualisiert"
        ws["A3"] = "Artikel"
        ws["B3"] = "Sollbestand"
        ws["C3"] = "AktuellerBestand"
        ws["D3"] = "Einheit"

        row = DATA_START_ROW
        for article, target, current, unit in group_data["items"]:
            ws.cell(row=row, column=1, value=article)
            ws.cell(row=row, column=2, value=target)
            ws.cell(row=row, column=3, value=current)
            ws.cell(row=row, column=4, value=unit)
            row += 1

    wb.save(path)


def ensure_workbook_exists() -> None:
    if not WORKBOOK_PATH.exists():
        create_default_workbook(WORKBOOK_PATH)


def get_sheet_or_none(wb, group_name: str):
    if group_name in wb.sheetnames:
        return wb[group_name]
    return None


def parse_group(ws) -> dict:
    supplier_email = str(ws["B1"].value or "").strip()
    items = []

    for row in range(DATA_START_ROW, ws.max_row + 1):
        article = ws.cell(row=row, column=1).value
        if article is None or str(article).strip() == "":
            continue

        target_stock = to_float(ws.cell(row=row, column=2).value)
        current_stock = to_float(ws.cell(row=row, column=3).value)
        unit = str(ws.cell(row=row, column=4).value or "Stk").strip()
        order_qty = max(target_stock - current_stock, 0.0)

        items.append(
            {
                "row": row,
                "article": str(article),
                "target_stock": target_stock,
                "current_stock": current_stock,
                "unit": unit,
                "order_qty": order_qty,
            }
        )

    items_to_order = sum(1 for item in items if item["order_qty"] > 0)

    return {
        "name": ws.title,
        "supplier_email": supplier_email,
        "items": items,
        "items_total": len(items),
        "items_to_order": items_to_order,
        "updated_at": ws["E1"].value,
    }


def update_current_stock(ws, updates: list[dict]) -> None:
    for update in updates:
        row = update.get("row")
        if not isinstance(row, int):
            continue
        if row < DATA_START_ROW or row > ws.max_row + 50:
            continue

        current_stock = to_float(update.get("current_stock"))
        ws.cell(row=row, column=3, value=current_stock)

    ws["E1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def build_mail_draft(group: dict) -> dict:
    receiver = group["supplier_email"]
    if not receiver:
        raise ValueError("Keine Lieferanten-E-Mail in der Excel fuer diese Lagergruppe hinterlegt.")

    to_order = [item for item in group["items"] if item["order_qty"] > 0]

    date_label = datetime.now().strftime("%d.%m.%Y")
    subject = f"Bestellung Lagergruppe {group['name']} ({date_label})"

    lines = [
        "Guten Tag,",
        "",
        f"bitte liefern Sie folgende Artikel fuer die Lagergruppe {group['name']}:",
        "",
    ]

    if not to_order:
        lines.append("- Aktuell keine Nachbestellung noetig.")
    else:
        for item in to_order:
            qty = format_number(item["order_qty"])
            current = format_number(item["current_stock"])
            target = format_number(item["target_stock"])
            lines.append(
                f"- {item['article']}: {qty} {item['unit']} (Bestand {current}, Soll {target})"
            )

    lines.extend(["", "Vielen Dank."])
    body = "\n".join(lines)
    mailto_url = f"mailto:{quote(receiver)}?subject={quote(subject)}&body={quote(body)}"

    return {
        "to": receiver,
        "subject": subject,
        "body": body,
        "mailto_url": mailto_url,
    }


@app.route("/")
def index():
    ensure_workbook_exists()
    return render_template("index.html")


@app.route("/api/groups", methods=["GET"])
def list_groups():
    ensure_workbook_exists()
    wb = load_workbook(WORKBOOK_PATH)

    groups = [parse_group(wb[sheet_name]) for sheet_name in wb.sheetnames]
    payload = [
        {
            "name": group["name"],
            "supplier_email": group["supplier_email"],
            "items_total": group["items_total"],
            "items_to_order": group["items_to_order"],
            "updated_at": group["updated_at"],
        }
        for group in groups
    ]
    return jsonify({"groups": payload})


@app.route("/api/groups/<group_name>", methods=["GET"])
def get_group(group_name: str):
    ensure_workbook_exists()
    wb = load_workbook(WORKBOOK_PATH)
    ws = get_sheet_or_none(wb, group_name)

    if ws is None:
        return jsonify({"error": "Lagergruppe nicht gefunden."}), 404

    return jsonify(parse_group(ws))


@app.route("/api/groups/<group_name>/save", methods=["POST"])
def save_group(group_name: str):
    ensure_workbook_exists()
    data = request.get_json(silent=True) or {}
    updates = data.get("items", [])

    wb = load_workbook(WORKBOOK_PATH)
    ws = get_sheet_or_none(wb, group_name)
    if ws is None:
        return jsonify({"error": "Lagergruppe nicht gefunden."}), 404

    if isinstance(updates, list):
        update_current_stock(ws, updates)

    wb.save(WORKBOOK_PATH)
    return jsonify({"ok": True, "group": parse_group(ws)})


@app.route("/api/groups/<group_name>/mail-draft", methods=["POST"])
def create_mail_draft(group_name: str):
    ensure_workbook_exists()
    data = request.get_json(silent=True) or {}
    updates = data.get("items", [])

    wb = load_workbook(WORKBOOK_PATH)
    ws = get_sheet_or_none(wb, group_name)
    if ws is None:
        return jsonify({"error": "Lagergruppe nicht gefunden."}), 404

    if isinstance(updates, list):
        update_current_stock(ws, updates)
        wb.save(WORKBOOK_PATH)

    group = parse_group(ws)

    try:
        mail = build_mail_draft(group)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    return jsonify({"ok": True, "group": group, "mail": mail})


if __name__ == "__main__":
    ensure_workbook_exists()
    app.run(host="127.0.0.1", port=5000, debug=True)
